#!/usr/bin/env python3
import argparse
import csv
import os
from collections import defaultdict


METRIC_COLUMNS = ['image_AUROC', 'image_AP', 'image_F1', 'normal_mean', 'abnormal_mean']


def _to_float(value):
    if value is None or value == '':
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _fmt_percent(value):
    if value is None:
        return ''
    return value * 100.0


def _read_rows(path):
    with open(path, 'r', newline='') as f:
        return list(csv.DictReader(f))


def _write_csv(path, rows, fieldnames):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _build_tables(rows):
    avg_rows = []
    organ_rows = []
    by_mode = defaultdict(dict)
    for row in rows:
        mode = row.get('score_mode', '')
        class_name = row.get('class_name', '')
        values = {
            'score_mode': mode,
            'score_description': row.get('score_description', ''),
            'class_name': class_name,
        }
        for col in METRIC_COLUMNS:
            values[col] = _to_float(row.get(col))
        by_mode[mode][class_name] = values
        if class_name == 'Avg':
            avg_rows.append(values)
        else:
            organ_rows.append(values)

    avg_rows.sort(
        key=lambda item: -1e9 if item['image_AUROC'] is None else item['image_AUROC'],
        reverse=True,
    )

    comparison_rows = []
    baseline = by_mode.get('model_top1', {}).get('Avg')
    baseline_auc = baseline.get('image_AUROC') if baseline else None
    for idx, row in enumerate(avg_rows, start=1):
        auc = row.get('image_AUROC')
        comparison_rows.append({
            'rank': idx,
            'score_mode': row['score_mode'],
            'score_description': row['score_description'],
            'Avg Image AUROC (%)': _fmt_percent(auc),
            'Delta vs model_top1 (%)': _fmt_percent(auc - baseline_auc)
            if auc is not None and baseline_auc is not None else '',
            'Avg Image AP (%)': _fmt_percent(row.get('image_AP')),
            'Avg Image F1 (%)': _fmt_percent(row.get('image_F1')),
            'normal_mean': row.get('normal_mean') if row.get('normal_mean') is not None else '',
            'abnormal_mean': row.get('abnormal_mean') if row.get('abnormal_mean') is not None else '',
        })

    organ_table = []
    for row in sorted(organ_rows, key=lambda item: (item['score_mode'], item['class_name'])):
        organ_table.append({
            'score_mode': row['score_mode'],
            'score_description': row['score_description'],
            'organ': row['class_name'],
            'Image AUROC (%)': _fmt_percent(row.get('image_AUROC')),
            'Image AP (%)': _fmt_percent(row.get('image_AP')),
            'Image F1 (%)': _fmt_percent(row.get('image_F1')),
            'normal_mean': row.get('normal_mean') if row.get('normal_mean') is not None else '',
            'abnormal_mean': row.get('abnormal_mean') if row.get('abnormal_mean') is not None else '',
        })
    return comparison_rows, organ_table


def _write_xlsx(path, comparison_rows, organ_rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = 'Avg Ranking'
    ws2 = wb.create_sheet('Per Organ')
    ws3 = wb.create_sheet('Notes')

    header_fill = PatternFill('solid', fgColor='17365D')
    header_font = Font(color='FFFFFF', bold=True)
    title_font = Font(bold=True, size=14)
    thin = Side(style='thin', color='D9E2F3')
    border = Border(bottom=thin)

    def write_table(sheet, title, rows):
        sheet['A1'] = title
        sheet['A1'].font = title_font
        if not rows:
            return
        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row_idx, row in enumerate(rows, start=4):
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=row[header])
                cell.border = border
                if isinstance(row[header], float):
                    cell.number_format = '0.00'
                if col_idx in (2, 3):
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        sheet.freeze_panes = 'A4'
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row_idx in range(4, min(sheet.max_row, 80) + 1):
                max_len = max(max_len, len(str(sheet.cell(row=row_idx, column=col_idx).value or '')))
            width = min(max(max_len + 2, 12), 55)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    write_table(ws, 'Map-to-Image Score Ranking', comparison_rows)
    write_table(ws2, 'Per-organ Map-to-Image Score Metrics', organ_rows)
    ws3['A1'] = 'Notes'
    ws3['A1'].font = title_font
    notes = [
        ('Goal', 'Compare image-level anomaly scores derived from the same pixel anomaly map.'),
        ('model_top1', 'Original model image_score baseline, usually whole-image top 1% mean.'),
        ('fg_top*', 'Foreground-only top-k pooling; reduces black background contamination.'),
        ('fg_eroded_top*', 'Foreground eroded before top-k; reduces organ edge false positives.'),
        ('fg_lse_t*', 'Normalized log-sum-exp pooling; soft-max-like but less sensitive to single noisy pixels.'),
        ('fg_struct_edge_penalty', 'Interior/foreground top-k score with foreground-edge penalty.'),
    ]
    for row_idx, (key, val) in enumerate(notes, start=3):
        ws3.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws3.cell(row=row_idx, column=2, value=val).alignment = Alignment(wrap_text=True, vertical='top')
    ws3.column_dimensions['A'].width = 24
    ws3.column_dimensions['B'].width = 90

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return True


def main():
    parser = argparse.ArgumentParser(description='Summarize map-to-image score variants from DebugEval CSVs.')
    parser.add_argument(
        '--debug-eval-dir',
        required=True,
        help='Path to a debug_eval directory containing foreground_score_sweep_metrics.csv.',
    )
    parser.add_argument(
        '--output-dir',
        default=None,
        help='Output directory. Default: <debug_eval_dir>/map_to_image_score_summary',
    )
    args = parser.parse_args()

    debug_eval_dir = args.debug_eval_dir
    source_csv = os.path.join(debug_eval_dir, 'foreground_score_sweep_metrics.csv')
    if not os.path.isfile(source_csv):
        raise FileNotFoundError(f'Missing {source_csv}. Run test/debug_eval first.')

    rows = _read_rows(source_csv)
    comparison_rows, organ_rows = _build_tables(rows)
    output_dir = args.output_dir or os.path.join(debug_eval_dir, 'map_to_image_score_summary')
    os.makedirs(output_dir, exist_ok=True)

    avg_csv = os.path.join(output_dir, 'map_to_image_score_avg_ranking.csv')
    organ_csv = os.path.join(output_dir, 'map_to_image_score_per_organ.csv')
    _write_csv(avg_csv, comparison_rows, list(comparison_rows[0].keys()) if comparison_rows else [])
    _write_csv(organ_csv, organ_rows, list(organ_rows[0].keys()) if organ_rows else [])

    xlsx_path = os.path.join(output_dir, 'map_to_image_score_summary.xlsx')
    wrote_xlsx = _write_xlsx(xlsx_path, comparison_rows, organ_rows)
    print(f'Wrote: {avg_csv}')
    print(f'Wrote: {organ_csv}')
    if wrote_xlsx:
        print(f'Wrote: {xlsx_path}')
    else:
        print('openpyxl is not installed; CSV files were written, XLSX skipped.')


if __name__ == '__main__':
    main()
